from typing import List
from datetime import datetime
from os.path import join

import flet as ft
from openpyxl import load_workbook, Workbook

from src.db.db import DBConn, SKU
from src.api.api import Playwright, BrowserConfig
from src.component.log_viewer import LogViewer
from src.component.table import Table
from src.component.crawler import Crawler

# 全局数据
db = DBConn()
playwright = Playwright(config=BrowserConfig(headless=False))

# 初始化数据库
db.create_tables()


# 导入需要处理的文件
async def handle_import_file(_: ft.Event[ft.Button]):
    """模拟数据导入"""
    # 这里可以放置实际的数据导入逻辑，例如从文件或数据库加载数据
    files = await ft.FilePicker().pick_files(
        dialog_title="选择文件进行导入",
        allowed_extensions=["xlsx"],
    )

    for file in files:
        wb = load_workbook(file.path)  # 这里可以根据需要处理 Excel 文件
        for ws in wb.worksheets:
            for row in ws.iter_rows(min_row=2, values_only=True):  # 从第几行开始
                if len(row) < 9:
                    continue  # 确保行数据足够

                origin_id = row[2]
                origin_shop_name = row[1]
                origin_primary_image_link = row[3]
                origin_price = row[6]
                origin_link = row[7]

                db.create_sku(
                    SKU(
                        origin_id=origin_id,
                        origin_shop_name=origin_shop_name,
                        origin_primary_image_link=origin_primary_image_link,
                        origin_link=origin_link,
                        origin_price=origin_price,
                    )
                )


async def handle_export_file(_: ft.Event[ft.Button]):
    # 选择保存路径
    directory_path = await ft.FilePicker().get_directory_path()

    wb = Workbook(write_only=True)
    ws = wb.create_sheet()

    # 写入表头
    ws.append(["店铺名称", "链接"])

    infos = db.query_sku()
    items: List[SKU] = []
    # 写入数据
    for row in infos:
        items.append(row.origin_id)
        # "店铺名称", "链接"
        ws.append((row.origin_shop_name, row.origin_link))

    # 保存
    now = datetime.now()
    filename = join(directory_path, f"蓝海词比价_{now.strftime('%m%d_%H%M%S')}.xlsx")
    wb.save(filename=filename)

    # 更新状态
    db.batch_modify_sku_export_status(items)


@ft.component
def App(playwright: Playwright) -> ft.Row:
    return ft.Row(
        expand=True,
        vertical_alignment=ft.CrossAxisAlignment.STRETCH,
        controls=[
            ft.Container(
                expand=8,
                content=ft.Column(
                    expand=True,
                    controls=[
                        ft.Row(
                            controls=[
                                ft.Button(
                                    "导入",
                                    icon=ft.Icons.UPLOAD_FILE,
                                    on_click=handle_import_file,
                                    color=ft.Colors.WHITE,
                                    bgcolor=ft.Colors.BLUE_600,
                                ),
                                ft.Button(
                                    "导出",
                                    icon=ft.Icons.DOWNLOAD,
                                    on_click=handle_export_file,
                                    color=ft.Colors.WHITE,
                                    bgcolor=ft.Colors.BLUE_600,
                                ),
                            ]
                        ),
                        ft.Divider(),
                        ft.Column(
                            expand=True,
                            controls=[
                                Table(
                                    # [
                                    #     SKU(
                                    #         origin_shop_name="店铺A",
                                    #         origin_primary_image_link="https://cbu01.alicdn.com/img/ibank/O1CN01Y6tpy129Wrl2dlnmX_!!2171508076-0-cib.jpg",
                                    #         origin_price="9.99",
                                    #         match_link="https://example.com/match",
                                    #         match_image_link="https://cbu01.alicdn.com/img/ibank/O1CN01Y6tpy129Wrl2dlnmX_!!2171508076-0-cib.jpg",
                                    #         match_score=0.85,
                                    #         status=1,
                                    #     ),
                                    #     SKU(
                                    #         origin_shop_name="店铺A",
                                    #         origin_primary_image_link="https://cbu01.alicdn.com/img/ibank/O1CN01Y6tpy129Wrl2dlnmX_!!2171508076-0-cib.jpg",
                                    #         origin_price="9.99",
                                    #         match_link="https://example.com/match",
                                    #         match_image_link="https://cbu01.alicdn.com/img/ibank/O1CN01Y6tpy129Wrl2dlnmX_!!2171508076-0-cib.jpg",
                                    #         match_score=0.85,
                                    #         status=1,
                                    #     ),
                                    #     SKU(
                                    #         origin_shop_name="店铺A",
                                    #         origin_primary_image_link="https://cbu01.alicdn.com/img/ibank/O1CN01Y6tpy129Wrl2dlnmX_!!2171508076-0-cib.jpg",
                                    #         origin_price="9.99",
                                    #         match_link="https://example.com/match",
                                    #         match_image_link="https://cbu01.alicdn.com/img/ibank/O1CN01Y6tpy129Wrl2dlnmX_!!2171508076-0-cib.jpg",
                                    #         match_score=0.85,
                                    #         status=1,
                                    #     ),
                                    #     SKU(
                                    #         origin_shop_name="店铺A",
                                    #         origin_primary_image_link="https://cbu01.alicdn.com/img/ibank/O1CN01Y6tpy129Wrl2dlnmX_!!2171508076-0-cib.jpg",
                                    #         origin_price="9.99",
                                    #         match_link="https://example.com/match",
                                    #         match_image_link="https://cbu01.alicdn.com/img/ibank/O1CN01Y6tpy129Wrl2dlnmX_!!2171508076-0-cib.jpg",
                                    #         match_score=0.85,
                                    #         status=1,
                                    #     ),
                                    #     SKU(
                                    #         origin_shop_name="店铺A",
                                    #         origin_primary_image_link="https://cbu01.alicdn.com/img/ibank/O1CN01Y6tpy129Wrl2dlnmX_!!2171508076-0-cib.jpg",
                                    #         origin_price="9.99",
                                    #         match_link="https://example.com/match",
                                    #         match_image_link="https://cbu01.alicdn.com/img/ibank/O1CN01Y6tpy129Wrl2dlnmX_!!2171508076-0-cib.jpg",
                                    #         match_score=0.85,
                                    #         status=1,
                                    #     ),
                                    # ]
                                    db.query_sku()
                                )
                            ],
                            scroll=ft.ScrollMode.AUTO,
                        ),
                    ],
                ),
            ),
            ft.Container(
                expand=2,
                content=ft.Column(
                    expand=True,
                    controls=[
                        ft.Container(
                            expand=3,
                            content=Crawler(db, playwright),
                        ),
                        ft.Container(
                            expand=7,
                            content=LogViewer(),
                        ),
                    ],
                ),
            ),
        ],
    )


async def main(page: ft.Page):
    page.title = "SKU 比价工具"

    page.window.width = 1100
    page.window.min_width = 1100
    page.window.min_height = 600

    # 初始化自动化环境
    await playwright.setup()
    page.render(lambda: App(playwright))


if __name__ == "__main__":
    ft.run(main)
